#!/usr/bin/env python3
"""
Excel Tool - Excel 操作工具
支持收据的保存、读取、查询等操作
"""

import json
import os
from pathlib import Path
from typing import Optional
from datetime import date
from decimal import Decimal
import logging

from receipt_manager.excel_handler import ExcelHandler
from receipt_manager import PurchaseReceipt
from copy import copy
import zipfile
import xml.etree.ElementTree as ET

logger = logging.getLogger(__name__)


def _get_sheet_order_from_xml(excel_file: Path) -> list:
    """
    直接从 Excel 文件的 XML 中读取 Sheet 的实际顺序
    这比 openpyxl.sheetnames 更可靠，因为它返回的是文件中存储的实际顺序
    """
    try:
        with zipfile.ZipFile(excel_file, 'r') as zip_ref:
            # 读取 workbook.xml
            workbook_xml = zip_ref.read('xl/workbook.xml')
            root = ET.fromstring(workbook_xml)

            # 命名空间
            ns = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

            # 获取所有 sheet 元素
            sheets = root.findall('.//main:sheet', ns)

            # 提取 sheet 名称（按文件中存储的顺序）
            sheet_names = []
            for sheet in sheets:
                name = sheet.get('name')
                if name:
                    sheet_names.append(name)

            return sheet_names
    except Exception as e:
        # 如果 XML 解析失败，回退到 openpyxl
        return None


def _normalize_path(path: str) -> Path:
    """
    规范化文件路径，处理各种常见格式

    Args:
        path: 输入路径，可能是：
            - ~/Downloads/file.xlsx
            - ~/home/user/Downloads/file.xlsx (错误格式)
            - /home/user/Downloads/file.xlsx (绝对路径)
            - ./file.xlsx (相对路径)

    Returns:
        规范化后的 Path 对象
    """
    # 处理 LLM 可能生成的错误格式: ~/home/user/...
    # 例如: ~/home/bughero/Documents/file.xlsx
    # 实际上应该是 /home/user/Downloads/file.xlsx
    if path.startswith("~/home/") or path.startswith("~/\\home\\"):
        # 去掉 ~/ 前缀
        actual_path = path[2:]  # 变成 home/user/...
        # 添加根目录 / 使其成为绝对路径
        return Path("/" + actual_path)

    p = Path(path)

    # 如果已经是绝对路径，直接返回
    if p.is_absolute():
        return p

    # 如果路径以 ~/ 开头（正常格式）
    if path.startswith("~/"):
        return Path(path).expanduser()

    # 相对路径
    return Path(path).expanduser()


def save_receipt_to_excel(
    receipt_data: dict,
    excel_path: str = "~/Documents/receipt-309/309-采购明细.xlsx",
    image_path: Optional[str] = None,
) -> str:
    """
    保存收据到 Excel 文件

    Args:
        receipt_data: 收据数据字典，包含 title, delivery_date, purchaser, payment_method, items
        excel_path: Excel 文件路径
        image_path: 收据图片路径（可选），支持 .jpg, .png, .bmp, .gif 等格式
                    图片将自动插入到表格内容下方

    Returns:
        操作结果的格式化字符串

    Example:
        save_receipt_to_excel({
            "title": "数学资料打印",
            "delivery_date": "2025-01-20",
            "purchaser": "梁程程妈妈",
            "payment_method": "转账",
            "items": [
                {"sequence": 1, "name": "语文资料", "quantity": 47, "unit_price": 6.0, "unit": "份"}
            ]
        }, image_path="./receipt.jpg")
    """
    try:
        # 解析日期
        delivery_date = receipt_data.get("delivery_date")
        if isinstance(delivery_date, str):
            delivery_date = date.fromisoformat(delivery_date)
        elif not isinstance(delivery_date, date):
            delivery_date = date.today()

        # 解析商品列表
        items_data = receipt_data.get("items", [])
        from receipt_manager import PurchaseItem, create_receipt

        items = []
        for item_data in items_data:
            item = PurchaseItem(
                sequence=item_data.get("sequence", len(items) + 1),
                name=item_data.get("name", ""),
                spec=item_data.get("spec"),
                unit=item_data.get("unit", "个"),
                quantity=Decimal(str(item_data.get("quantity", 1))),
                unit_price=Decimal(str(item_data.get("unit_price", 0))),
                remark=item_data.get("remark"),
            )
            items.append(item)

        # 创建收据（传递图片路径给 source_file，让 ExcelHandler 正确插入到表格下方）
        receipt = create_receipt(
            title=receipt_data.get("title", "未命名采购"),
            delivery_date=delivery_date,
            purchaser=receipt_data.get("purchaser", "梁程程妈妈"),
            source_file=image_path,  # 图片路径将通过 source_file 传递
        )
        receipt.payment_method = receipt_data.get("payment_method", "转账")
        receipt.items = items

        # 保存到 Excel（图片通过 source_file 由 ExcelHandler 统一处理，插入到表格下方）
        excel_file = _normalize_path(excel_path)
        handler = ExcelHandler(excel_file)
        handler.add_receipt(receipt)
        handler.close()

        # 图片信息（由 ExcelHandler 处理）
        image_info = ""
        if image_path:
            img_file = _normalize_path(image_path)
            if img_file.exists():
                image_info = (
                    f"\n  📷 图片已插入\n"
                    f"    位置: 表格下方\n"
                    f"    文件: {img_file.name}"
                )
            else:
                image_info = f"\n  ⚠ 图片文件不存在: {image_path}"

        return (
            f"✓ 收据已保存到 Excel\n"
            f"  文件: {excel_file}\n"
            f"  Sheet: {receipt.sheet_name}\n"
            f"  主题: {receipt.title}\n"
            f"  金额: ¥{receipt.total_amount:.2f}"
            f"{image_info}"
        )

    except Exception as e:
        return f"❌ 保存失败: {str(e)}"


def read_receipt_from_excel(
    sheet_name: str,
    excel_path: str = "~/Documents/receipt-309/309-采购明细.xlsx",
) -> str:
    """
    从 Excel 读取收据

    Args:
        sheet_name: Sheet 名称
        excel_path: Excel 文件路径

    Returns:
        收据数据的格式化字符串

    Example:
        read_receipt_from_excel(sheet_name="数学资料打印（1-20）")
    """
    try:
        excel_file = _normalize_path(excel_path)
        handler = ExcelHandler(excel_file)

        receipt = handler.read_receipt(sheet_name)
        handler.close()

        if receipt is None:
            return f"❌ 未找到 Sheet: {sheet_name}"

        # 格式化输出
        lines = [
            f"📄 收据信息",
            f"{'='*60}",
            f"  Sheet: {sheet_name}",
            f"  主题: {receipt.title}",
            f"  日期: {receipt.delivery_date}",
            f"  采购方: {receipt.purchaser}",
            f"  付款方式: {receipt.payment_method}",
            f"{'='*60}",
            f"{'序号':<6} {'商品名称':<25} {'单位':<6} {'数量':>8} {'单价':>10} {'金额':>10}",
            f"{'-'*60}",
        ]

        for item in receipt.items:
            name = item.name[:25]
            lines.append(
                f"{item.sequence:<6} {name:<25} {item.unit:<6} "
                f"{float(item.quantity):>8.1f} ¥{float(item.unit_price):>9.2f} ¥{float(item.amount):>9.2f}"
            )

        lines.extend([
            f"{'-'*60}",
            f"{'':>38} {'总计':>10} ¥{float(receipt.total_amount):>9.2f}",
            f"{'='*60}",
            f"  商品数量: {receipt.item_count}",
            f"  总数量: {float(receipt.total_quantity):.1f}",
        ])

        return "\n".join(lines)

    except Exception as e:
        return f"❌ 读取失败: {str(e)}"


def list_excel_sheets(
    excel_path: str = "~/Documents/receipt-309/309-采购明细.xlsx",
) -> str:
    """
    列出 Excel 中所有 Sheet

    Args:
        excel_path: Excel 文件路径

    Returns:
        Sheet 列表的格式化字符串

    Example:
        list_excel_sheets()
    """
    try:
        excel_file = _normalize_path(excel_path)

        if not excel_file.exists():
            return f"❌ Excel 文件不存在: {excel_file}"

        # 直接使用 openpyxl 读取，确保顺序正确
        # 注意：不能使用 read_only=True，因为它会按字母顺序返回 sheets
        from openpyxl import load_workbook
        wb = load_workbook(excel_file)
        sheets = wb.sheetnames
        wb.close()

        if not sheets:
            return "📭 Excel 中没有 Sheet"

        lines = [
            f"📋 Excel Sheet 列表",
            f"{'='*60}",
            f"  文件: {excel_path}",
            f"{'='*60}",
        ]

        for i, sheet in enumerate(sheets, 1):
            lines.append(f"  {i}. {sheet}")

        lines.append(f"{'='*60}")
        lines.append(f"  总计: {len(sheets)} 个 Sheet")

        return "\n".join(lines)

    except Exception as e:
        return f"❌ 列出失败: {str(e)}"


def update_receipt_in_excel(
    sheet_name: str,
    receipt_data: dict,
    excel_path: str = "~/Documents/receipt-309/309-采购明细.xlsx",
    image_path: Optional[str] = None,
) -> str:
    """
    更新 Excel 中的收据

    Args:
        sheet_name: Sheet 名称
        receipt_data: 新的收据数据
        excel_path: Excel 文件路径
        image_path: 收据图片路径（可选），支持 .jpg, .png, .bmp, .gif 等格式
                    图片将自动插入到表格内容下方

    Returns:
        操作结果的格式化字符串

    Example:
        update_receipt_in_excel(
            sheet_name="数学资料打印（1-20）",
            receipt_data={"title": "新标题", ...},
            image_path="./receipt.jpg"
        )
    """
    try:
        excel_file = _normalize_path(excel_path)
        handler = ExcelHandler(excel_file)

        # 检查 Sheet 是否存在
        if not handler.sheet_exists(sheet_name):
            handler.close()
            return f"❌ Sheet 不存在: {sheet_name}"

        # 解析并保存
        # 先读取现有收据获取默认值
        existing = handler.read_receipt(sheet_name)

        # 解析日期
        delivery_date = receipt_data.get("delivery_date")
        if isinstance(delivery_date, str):
            delivery_date = date.fromisoformat(delivery_date)
        elif delivery_date is None and existing:
            delivery_date = existing.delivery_date
        else:
            delivery_date = date.today()

        # 解析商品列表
        from receipt_manager import PurchaseItem, create_receipt

        items = []
        items_data = receipt_data.get("items", [])
        if items_data:
            for item_data in items_data:
                item = PurchaseItem(
                    sequence=item_data.get("sequence", len(items) + 1),
                    name=item_data.get("name", ""),
                    spec=item_data.get("spec"),
                    unit=item_data.get("unit", "个"),
                    quantity=Decimal(str(item_data.get("quantity", 1))),
                    unit_price=Decimal(str(item_data.get("unit_price", 0))),
                    remark=item_data.get("remark"),
                )
                items.append(item)
        elif existing:
            items = existing.items

        # 创建收据（传递图片路径给 source_file，让 ExcelHandler 正确插入到表格下方）
        receipt = create_receipt(
            title=receipt_data.get("title", existing.title if existing else "未命名"),
            delivery_date=delivery_date,
            purchaser=receipt_data.get("purchaser", existing.purchaser if existing else "梁程程妈妈"),
            source_file=image_path,  # 图片路径将通过 source_file 传递
        )
        receipt.payment_method = receipt_data.get("payment_method", existing.payment_method if existing else "转账")
        receipt.items = items

        handler.add_receipt(receipt)
        handler.close()

        # 图片信息（由 ExcelHandler 处理）
        image_info = ""
        if image_path:
            img_file = _normalize_path(image_path)
            if img_file.exists():
                image_info = (
                    f"\n  📷 图片已插入\n"
                    f"    位置: 表格下方\n"
                    f"    文件: {img_file.name}"
                )
            else:
                image_info = f"\n  ⚠ 图片文件不存在: {image_path}"

        return (
            f"✓ 收据已更新\n"
            f"  Sheet: {sheet_name}\n"
            f"  主题: {receipt.title}\n"
            f"  金额: ¥{receipt.total_amount:.2f}"
            f"{image_info}"
        )

    except Exception as e:
        return f"❌ 更新失败: {str(e)}"


def delete_receipt_from_excel(
    sheet_name: str,
    excel_path: str = "~/Documents/receipt-309/309-采购明细.xlsx",
) -> str:
    """
    删除 Excel 中的收据 Sheet

    Args:
        sheet_name: Sheet 名称
        excel_path: Excel 文件路径

    Returns:
        操作结果的格式化字符串

    Example:
        delete_receipt_from_excel(sheet_name="数学资料打印（1-20）")
    """
    try:
        from openpyxl import load_workbook

        excel_file = _normalize_path(excel_path)

        if not excel_file.exists():
            return f"❌ Excel 文件不存在: {excel_file}"

        wb = load_workbook(excel_file)

        if sheet_name not in wb.sheetnames:
            return f"❌ Sheet 不存在: {sheet_name}"

        wb.remove(wb[sheet_name])
        wb.save(excel_file)

        return f"✓ 已删除 Sheet: {sheet_name}"

    except Exception as e:
        return f"❌ 删除失败: {str(e)}"


def merge_excel_files(
    source_paths: list,
    target_path: str = "~/Documents/receipt-309/309-采购明细.xlsx",
    remove_duplicates: bool = True,
) -> str:
    """
    合并多个 Excel 文件到目标文件

    Args:
        source_paths: 源 Excel 文件路径列表
        target_path: 目标 Excel 文件路径
        remove_duplicates: 是否删除重复的 Sheet（按 Sheet 名称）

    Returns:
        操作结果的格式化字符串

    Example:
        merge_excel_files(source_paths=["./file1.xlsx", "./file2.xlsx"])
    """
    try:
        from openpyxl import load_workbook
        import xlrd  # 用于处理 .xls 格式

        target_file = _normalize_path(target_path)

        # 确保目标文件存在，如果不存在则创建
        if not target_file.exists():
            from openpyxl import Workbook
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            wb.save(target_file)

        # 加载目标工作簿
        target_wb = load_workbook(target_file)
        existing_sheets = set(target_wb.sheetnames)

        merged_count = 0
        skipped_count = 0
        duplicate_count = 0

        for source_path_str in source_paths:
            source_path = _normalize_path(source_path_str)
            if not source_path.exists():
                return f"❌ 源文件不存在: {source_path_str}"

            # 如果是同一个文件，跳过
            if source_path == target_file:
                continue

            try:
                # 检测文件格式
                # 首先尝试通过扩展名检测
                is_xls = source_path.suffix.lower() == '.xls'

                # 如果扩展名检测失败，尝试通过文件头 magic bytes 检测
                if not is_xls and source_path.exists():
                    try:
                        with open(source_path, 'rb') as f:
                            header = f.read(8)
                            # .xls 文件的 magic bytes (OLE2 compound document)
                            # D0 CF 11 E0 A1 B1 1A E1
                            if header[:8] == b'\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1':
                                is_xls = True
                    except Exception:
                        pass

                if is_xls:
                    # 使用 xlrd 读取 .xls 文件
                    try:
                        source_wb = xlrd.open_workbook(str(source_path))

                        for sheet_idx in range(source_wb.nsheets):
                            sheet_name = source_wb.sheet_names()[sheet_idx]

                            # 检查是否重复
                            if sheet_name in existing_sheets:
                                if remove_duplicates:
                                    target_wb.remove(target_wb[sheet_name])
                                    duplicate_count += 1
                                else:
                                    skipped_count += 1
                                    continue

                            # 创建新 sheet
                            from openpyxl import Workbook
                            target_ws = target_wb.create_sheet(title=sheet_name)

                            # 读取 .xls 数据并写入
                            source_sheet = source_wb.sheet_by_index(sheet_idx)
                            for row_idx in range(source_sheet.nrows):
                                for col_idx in range(source_sheet.ncols):
                                    cell = source_sheet.cell(row_idx, col_idx)
                                    target_ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell.value)

                            merged_count += 1
                            existing_sheets.add(sheet_name)

                        source_wb.release_resources()
                    except Exception as e:
                        return f"❌ 读取 .xls 文件失败 {source_path.name}: {str(e)}"
                else:
                    # 使用 openpyxl 读取 .xlsx 文件
                    source_wb = load_workbook(source_path)

                    for sheet_name in source_wb.sheetnames:
                        # 检查是否重复
                        if sheet_name in existing_sheets:
                            if remove_duplicates:
                                # 删除目标中的旧 sheet
                                target_wb.remove(target_wb[sheet_name])
                                duplicate_count += 1
                            else:
                                skipped_count += 1
                                continue

                        # 复制 sheet
                        source_ws = source_wb[sheet_name]
                        target_ws = target_wb.create_sheet(title=sheet_name)

                        # 复制所有单元格（保留原始值和格式）
                        for row in source_ws.iter_rows():
                            for cell in row:
                                target_cell = target_ws.cell(row=cell.row, column=cell.column)

                                # 复制原始值（保留 Excel 日期序列号等）
                                target_cell.value = cell.value
                                if cell.data_type:
                                    target_cell.data_type = cell.data_type

                                # 复制样式
                                if cell.has_style:
                                    target_cell.font = copy(cell.font)
                                    target_cell.border = copy(cell.border)
                                    target_cell.fill = copy(cell.fill)
                                    target_cell.number_format = cell.number_format
                                    target_cell.alignment = copy(cell.alignment)

                        # 复制行高
                        for row_idx, row_dim in source_ws.row_dimensions.items():
                            target_ws.row_dimensions[row_idx].height = row_dim.height

                        # 复制列宽
                        for col_idx, col_dim in source_ws.column_dimensions.items():
                            target_ws.column_dimensions[col_idx].width = col_dim.width

                        # 复制合并单元格
                        if hasattr(source_ws, 'merged_cells'):
                            for merged_range in list(source_ws.merged_cells.ranges):
                                target_ws.merge_cells(str(merged_range))

                        # 复制图片
                        if hasattr(source_ws, '_images'):
                            for img in source_ws._images:
                                target_ws.add_image(img)

                        merged_count += 1
                        existing_sheets.add(sheet_name)

                    source_wb.close()

            except Exception as e:
                return f"❌ 合并 {source_path.name} 失败: {str(e)}"

        # 保存目标文件
        target_wb.save(target_file)
        target_wb.close()

        lines = [
            f"✓ Excel 文件合并完成",
            f"  目标文件: {target_file}",
            f"  合并数量: {merged_count} 个 Sheet",
        ]
        if duplicate_count > 0:
            lines.append(f"  替换重复: {duplicate_count} 个 Sheet")
        if skipped_count > 0:
            lines.append(f"  跳过重复: {skipped_count} 个 Sheet")

        return "\n".join(lines)

    except Exception as e:
        return f"❌ 合并失败: {str(e)}"


def sort_sheets_by_date(
    excel_path: str = "~/Documents/receipt-309/309-采购明细.xlsx",
    order: str = "desc",
) -> str:
    """
    按 Sheet 名称中的日期对 Sheet 进行排序

    Args:
        excel_path: Excel 文件路径
        order: 排序顺序，"desc" 表示从新到旧，"asc" 表示从旧到新

    Returns:
        操作结果的格式化字符串

    Example:
        sort_sheets_by_date()  # 从新到旧排序
        sort_sheets_by_date(order="asc")  # 从旧到新排序
    """
    try:
        from openpyxl import load_workbook
        from datetime import datetime
        import re

        excel_file = _normalize_path(excel_path)

        if not excel_file.exists():
            return f"❌ Excel 文件不存在: {excel_file}"

        wb = load_workbook(excel_file)

        # 尝试从 XML 中获取实际的 sheet 顺序（更可靠）
        xml_sheet_order = _get_sheet_order_from_xml(excel_file)
        if xml_sheet_order:
            sheet_names = xml_sheet_order
        else:
            # 回退到 openpyxl 的方法
            sheet_names = wb.sheetnames

        # 提取日期的函数
        def extract_date(sheet_name):
            # 匹配格式: "主题名称（月-日）" 或 "主题名称（年-月-日）"
            # 同时支持全角括号（）和半角括号()
            match = re.search(r'[()（）](\d{1,4})-(\d{1,2})-(\d{1,2})[()（）]', sheet_name)
            if match:
                year, month, day = match.groups()
                if len(year) == 4:
                    return datetime(int(year), int(month), int(day))
                else:
                    # 如果只有月-日，假设是当年
                    current_year = datetime.now().year
                    return datetime(current_year, int(month), int(day))
            return datetime.min

        # 排序
        sorted_names = sorted(sheet_names, key=extract_date, reverse=(order == "desc"))

        # 调试：显示排序前后的对比
        def get_date_str(name):
            m = re.search(r'[()（）](\d{1,4})-(\d{1,2})-(\d{1,2})[()（）]', name)
            return m.group(0) if m else "无日期"

        debug_info = []
        debug_info.append(f"  原始顺序（前5个）:")
        for name in sheet_names[:5]:
            debug_info.append(f"    {get_date_str(name)} - {name[:30]}")
        debug_info.append(f"  排序后顺序（前5个）:")
        for name in sorted_names[:5]:
            debug_info.append(f"    {get_date_str(name)} - {name[:30]}")

        # 创建临时文件
        import tempfile
        temp_fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
        import os
        os.close(temp_fd)

        try:
            # 创建新的工作簿来重新排序
            from openpyxl import Workbook
            new_wb = Workbook()
            if "Sheet" in new_wb.sheetnames:
                new_wb.remove(new_wb["Sheet"])

            for sheet_name in sorted_names:
                ws = wb[sheet_name]
                new_ws = new_wb.create_sheet(title=sheet_name)

                # 复制所有内容
                for row in ws.iter_rows():
                    for cell in row:
                        new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                        if cell.has_style:
                            from copy import copy
                            new_cell.font = copy(cell.font)
                            new_cell.border = copy(cell.border)
                            new_cell.fill = copy(cell.fill)
                            new_cell.number_format = cell.number_format
                            new_cell.alignment = copy(cell.alignment)

                # 复制行高
                for row_idx, row_dim in ws.row_dimensions.items():
                    new_ws.row_dimensions[row_idx].height = row_dim.height

                # 复制列宽
                for col_idx, col_dim in ws.column_dimensions.items():
                    new_ws.column_dimensions[col_idx].width = col_dim.width

                # 复制合并单元格
                if hasattr(ws, 'merged_cells'):
                    for merged_range in list(ws.merged_cells.ranges):
                        new_ws.merge_cells(str(merged_range))

                # 复制图片
                if hasattr(ws, '_images'):
                    for img in ws._images:
                        new_ws.add_image(img)

            # 先保存到临时文件
            new_wb.save(temp_path)
            new_wb.close()
            wb.close()

            # 用临时文件替换原文件
            import shutil
            shutil.move(temp_path, str(excel_file))

            # 验证排序结果
            verify_wb = load_workbook(excel_file)
            actual_order = verify_wb.sheetnames
            verify_wb.close()

            # 检查前5个和后5个的日期顺序
            def get_date_str(name):
                m = re.search(r'[()（）](\d{1,4})-(\d{1,2})-(\d{1,2})[()（）]', name)
                return m.group(0) if m else "无日期"

            preview = []
            for name in actual_order[:5]:
                preview.append(f"  {get_date_str(name)} - {name[:30]}")
            if len(actual_order) > 10:
                preview.append(f"  ... (共 {len(actual_order)} 个)")
            for name in actual_order[-5:]:
                preview.append(f"  {get_date_str(name)} - {name[:30]}")

            return (
                f"✓ Sheet 已按日期{'从新到旧' if order == 'desc' else '从旧到新'}排序\n"
                f"  总计: {len(sheet_names)} 个 Sheet\n"
                f"  验证结果:\n" + "\n".join(debug_info) + "\n"
                f"  文件保存后顺序（前5个）:\n" + "\n".join(preview)
            )

        except Exception as save_error:
            # 清理临时文件
            if os.path.exists(temp_path):
                os.unlink(temp_path)
            raise save_error

    except Exception as e:
        return f"❌ 排序失败: {str(e)}"


def rename_sheet(
    old_name: str,
    new_name: str,
    excel_path: str = "~/Documents/receipt-309/309-采购明细.xlsx",
) -> str:
    """
    重命名 Excel 中的 Sheet

    Args:
        old_name: 旧的 Sheet 名称
        new_name: 新的 Sheet 名称
        excel_path: Excel 文件路径

    Returns:
        操作结果的格式化字符串

    Example:
        rename_sheet(old_name="9月份开学教室布置", new_name="9月份开学教室布置（9-1）")
    """
    try:
        from openpyxl import load_workbook

        excel_file = _normalize_path(excel_path)

        if not excel_file.exists():
            return f"❌ Excel 文件不存在: {excel_file}"

        wb = load_workbook(excel_file)

        if old_name not in wb.sheetnames:
            return f"❌ Sheet 不存在: {old_name}"

        if new_name in wb.sheetnames:
            return f"❌ 目标名称已存在: {new_name}"

        # 重命名 Sheet
        ws = wb[old_name]
        ws.title = new_name
        wb.save(excel_file)
        wb.close()

        return f"✓ Sheet 已重命名: {old_name} → {new_name}"

    except Exception as e:
        return f"❌ 重命名失败: {str(e)}"


def rename_sheet_auto(
    old_name: str,
    date_str: str,
    excel_path: str = "~/Documents/receipt-309/309-采购明细.xlsx",
) -> str:
    """
    自动重命名 Sheet 为"主题（日期）"格式

    Args:
        old_name: 旧的 Sheet 名称
        date_str: 日期字符串，格式如 "9-1" 或 "2024-9-1"
        excel_path: Excel 文件路径

    Returns:
        操作结果的格式化字符串

    Example:
        rename_sheet_auto(old_name="9月份开学教室布置", date_str="9-1")
        # 结果: "9月份开学教室布置（9-1）"
    """
    try:
        from openpyxl import load_workbook

        excel_file = _normalize_path(excel_path)

        if not excel_file.exists():
            return f"❌ Excel 文件不存在: {excel_file}"

        wb = load_workbook(excel_file)

        if old_name not in wb.sheetnames:
            return f"❌ Sheet 不存在: {old_name}"

        # 检查旧名称是否已包含日期格式
        if "(" in old_name and ")" in old_name:
            return f"ℹ Sheet 名称已包含日期格式: {old_name}"

        # 生成新名称
        new_name = f"{old_name}（{date_str}）"

        if new_name in wb.sheetnames:
            return f"❌ 目标名称已存在: {new_name}"

        # 重命名 Sheet
        ws = wb[old_name]
        ws.title = new_name
        wb.save(excel_file)
        wb.close()

        return f"✓ Sheet 已重命名: {old_name} → {new_name}"

    except Exception as e:
        return f"❌ 重命名失败: {str(e)}"


def beautify_excel(
    excel_path: str = "~/Documents/receipt-309/309-采购明细.xlsx",
) -> str:
    """
    美化 Excel 文件 - 应用最新的专业样式到所有收据

    这个函数会重新格式化 Excel 文件中的所有收据 Sheet，应用最新的美化样式：
    - 深蓝色表头背景，白色文字
    - 浅黄色总计行，深红色文字
    - 浅灰色序号列背景
    - 微软雅黑字体，更清晰易读
    - 合适的行高和列宽
    - 专业的边框样式

    Args:
        excel_path: Excel 文件路径

    Returns:
        操作结果的格式化字符串

    Example:
        beautify_excel()
        beautify_excel(excel_path="~/Documents/收据.xlsx")
    """
    try:
        excel_file = _normalize_path(excel_path)

        if not excel_file.exists():
            return f"❌ Excel 文件不存在: {excel_file}"

        handler = ExcelHandler(excel_file)

        # 列出所有 Sheet
        sheet_names = handler.list_sheets()

        if not sheet_names:
            handler.close()
            return "📭 Excel 中没有 Sheet"

        # 美化所有 Sheet
        count = handler.beautify_all_sheets()
        handler.close()

        return (
            f"✓ Excel 美化完成\n"
            f"  文件: {excel_file}\n"
            f"  美化数量: {count} 个 Sheet\n"
            f"  应用样式:\n"
            f"    • 深蓝色表头背景 + 白色文字\n"
            f"    • 浅黄色总计行 + 深红色文字\n"
            f"    • 浅灰色序号/单位列背景\n"
            f"    • 微软雅黑字体（更清晰）\n"
            f"    • 专业的边框和间距"
        )

    except Exception as e:
        return f"❌ 美化失败: {str(e)}"


def merge_by_date(
    target_date: str,
    excel_path: str = "~/Documents/receipt-309/309-采购明细.xlsx",
    keep_original: bool = False,
) -> str:
    """
    按日期合并收据

    将指定日期的所有收据合并成一张汇总表，删除原 Sheet。

    Args:
        target_date: 目标日期，格式如 "2026-2-26" 或 "2-26"
        excel_path: Excel 文件路径
        keep_original: 是否保留原 Sheet（默认 False，即删除原 Sheet）

    Returns:
        操作结果的格式化字符串

    Example:
        merge_by_date("2026-2-26")
        merge_by_date("2-26")
    """
    try:
        from openpyxl import load_workbook

        excel_file = _normalize_path(excel_path)

        if not excel_file.exists():
            return f"❌ Excel 文件不存在: {excel_file}"

        wb = load_workbook(str(excel_file))
        handler = ExcelHandler(str(excel_file))

        # 解析日期（支持 "2026-2-26" 或 "2-26" 格式）
        import re
        date_match = re.match(r'(\d{4})?-(\d{1,2})-(\d{1,2})', target_date)
        if not date_match:
            return f"❌ 日期格式错误: {target_date}，请使用 YYYY-M-D 或 M-D 格式"

        year, month, day = date_match.groups()
        if year is None:
            year = date.today().year

        # 标准化日期（补零）
        target_year = int(year)
        target_month = int(month)
        target_day = int(day)

        # 查找匹配日期的 Sheet（排除已有的汇总表）
        matched_sheets = []
        for sheet_name in wb.sheetnames:
            # 排除包含"汇总"的 Sheet，避免重复合并
            if "汇总" in sheet_name:
                continue
            receipt = handler.read_receipt(sheet_name)
            if receipt and receipt.delivery_date:
                if (receipt.delivery_date.year == target_year and
                    receipt.delivery_date.month == target_month and
                    receipt.delivery_date.day == target_day):
                    matched_sheets.append((sheet_name, receipt))

        handler.close()

        if not matched_sheets:
            return f"❌ 未找到日期为 {target_year}-{target_month:02d}-{target_day:02d} 的收据"

        if len(matched_sheets) == 1:
            return f"ℹ 该日期只有 1 张收据，无需合并\n  Sheet: {matched_sheets[0][0]}"

        # 收集所有商品明细和图片
        all_items = []
        total_amount = 0
        purchaser = "梁程程妈妈"
        payment_method = "转账"
        titles = []
        sheet_images = []  # 收集每个 Sheet 的图片对象

        for sheet_name, receipt in matched_sheets:
            titles.append(receipt.title)
            if receipt.purchaser:
                purchaser = receipt.purchaser
            if receipt.payment_method:
                payment_method = receipt.payment_method

            # 从 Sheet 中提取图片对象
            ws = wb[sheet_name]
            if hasattr(ws, '_images') and ws._images:
                sheet_images.extend(ws._images)

            for item in receipt.items:
                all_items.append({
                    "sequence": len(all_items) + 1,
                    "name": item.name,
                    "spec": item.spec,
                    "unit": item.unit,
                    "quantity": float(item.quantity),
                    "unit_price": float(item.unit_price),
                    "remark": item.remark,
                })
            total_amount += float(receipt.total_amount)

        # 创建汇总收据数据（使用日期作为标题）
        summary_date = date(target_year, target_month, target_day)
        summary_title = f"{target_year}-{target_month}-{target_day}"  # 年-月-日格式
        receipt_data = {
            "title": summary_title,
            "delivery_date": summary_date.isoformat(),
            "purchaser": purchaser,
            "payment_method": payment_method,
            "items": all_items,
        }

        # 先保存汇总收据（不含图片，因为图片是从 Sheet 复制的）
        save_receipt_to_excel(
            receipt_data=receipt_data,
            excel_path=str(excel_file),
            image_path=None,  # 不通过 source_file 传图片
        )

        # 重新加载工作簿，获取汇总 Sheet 名称并插入图片
        wb = load_workbook(str(excel_file))
        summary_sheet_name = None
        for name in wb.sheetnames:
            if name.startswith(summary_title):
                summary_sheet_name = name
                break

        if summary_sheet_name and sheet_images:
            try:
                ws = wb[summary_sheet_name]

                # 找到表格最后一行（交付信息行之后）
                last_row = ws.max_row
                # 图片起始行（在表格内容之后）
                image_row = last_row + 2

                # 插入所有图片（横向排列）
                for i, img in enumerate(sheet_images):
                    try:
                        # 设置图片大小（如果图片太大）
                        max_width = 400  # 稍小一点以容纳多张图片
                        if img.width > max_width:
                            ratio = max_width / img.width
                            # 创建新的图片对象（避免修改原图片）
                            from copy import deepcopy
                            new_img = deepcopy(img)
                            new_img.width = max_width
                            new_img.height = int(img.height * ratio)
                        else:
                            new_img = img

                        # 计算图片位置（横向排列，每行2张）
                        col_offset = (i % 2) * 6  # 间隔6列
                        row_offset = (i // 2) * 25  # 每25行换行

                        target_col = 1 + col_offset  # A列或G列
                        target_row = image_row + row_offset

                        new_img.anchor = f"{chr(64 + target_col)}{target_row}"
                        ws.add_image(new_img)

                    except Exception as img_e:
                        logger.warning(f"插入第{i+1}张图片失败: {img_e}")

                wb.save(str(excel_file))
            except Exception as e:
                logger.warning(f"插入图片失败: {e}")
            finally:
                wb.close()

        # 重命名汇总表，去掉重复的日期后缀
        wb = load_workbook(str(excel_file))
        for name in wb.sheetnames:
            if name.startswith(summary_title):
                # 重命名为 "年-月-日" 格式，去掉多余的后缀
                ws = wb[name]
                ws.title = summary_title
                break
        wb.save(str(excel_file))
        wb.close()

        # 删除原 Sheet
        if not keep_original:
            wb = load_workbook(str(excel_file))
            deleted_count = 0
            for sheet_name, _ in matched_sheets:
                if sheet_name in wb.sheetnames:
                    wb.remove(wb[sheet_name])
                    deleted_count += 1
            wb.save(str(excel_file))
            wb.close()

            action = "已合并并删除"
        else:
            deleted_count = 0
            action = "已合并（保留原表）"

        # 格式化输出
        lines = [
            f"✓ 按日期合并完成",
            f"  目标日期: {target_year}-{target_month:02d}-{target_day:02d}",
            f"  合并数量: {len(matched_sheets)} 张收据",
            f"  {action}: {deleted_count} 张原 Sheet",
            f"  新 Sheet: {summary_title}",
            f"  商品数量: {len(all_items)} 件",
            f"  总金额: ¥{total_amount:.2f}",
            f"",
            f"  来源收据:",
        ]
        for sheet_name, receipt in matched_sheets:
            lines.append(f"    • {sheet_name}: ¥{float(receipt.total_amount):.2f}")

        return "\n".join(lines)

    except Exception as e:
        import traceback
        return f"❌ 合并失败: {str(e)}\n{traceback.format_exc()}"


# 导出所有工具函数
__all__ = [
    "save_receipt_to_excel",
    "read_receipt_from_excel",
    "list_excel_sheets",
    "update_receipt_in_excel",
    "delete_receipt_from_excel",
    "merge_excel_files",
    "sort_sheets_by_date",
    "rename_sheet",
    "rename_sheet_auto",
    "beautify_excel",
    "merge_by_date",
]

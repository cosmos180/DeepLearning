"""
Excel处理模块

负责读取和写入采购明细Excel文件，格式符合实际使用的309 采购明细.xlsx结构。
"""

from pathlib import Path
from typing import List, Optional
from datetime import date, datetime, timedelta
from decimal import Decimal
import shutil
import re
import logging

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.drawing.image import Image
except ImportError:
    raise ImportError(
        "需要安装 openpyxl: pip install openpyxl"
    )


# 样式定义
class Styles:
    """
    Excel样式定义 - 业界最佳实践

    设计原则：
    1. 层次分明：标题 > 表头 > 总计 > 数据
    2. 色彩和谐：使用柔和的配色方案，避免过于鲜艳
    3. 易于阅读：合适的字体大小、行高、列宽
    4. 专业规范：符合财务报表的行业标准
    """

    # ========== 边框样式 ==========
    # 细边框（用于数据单元格）
    THIN_BORDER = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )

    # 表头边框（稍粗，用于表头）
    HEADER_BORDER = Border(
        left=Side(style='medium', color='A0A0A0'),
        right=Side(style='medium', color='A0A0A0'),
        top=Side(style='medium', color='A0A0A0'),
        bottom=Side(style='medium', color='A0A0A0'),
    )

    # 总计边框（双线下边框）
    TOTAL_BORDER = Border(
        left=Side(style='thin', color='A0A0A0'),
        right=Side(style='thin', color='A0A0A0'),
        top=Side(style='thin', color='A0A0A0'),
        bottom=Side(style='double', color='404040'),
    )

    # ========== 字体样式 ==========
    # 标题字体（更大、更醒目）
    TITLE_FONT = Font(
        size=18,
        bold=True,
        color='2C3E50',  # 深蓝灰色，更专业
        name='微软雅黑',  # 使用微软雅黑字体
    )

    # 表头字体（加粗、深色）
    HEADER_FONT = Font(
        size=12,
        bold=True,
        color='FFFFFF',  # 白色文字，配合深色背景
        name='微软雅黑',
    )

    # 数据行字体（常规、深灰）
    DATA_FONT = Font(
        size=11,
        color='34495E',  # 深灰色，易于阅读
        name='微软雅黑',
    )

    # 总计字体（加粗、突出）
    TOTAL_FONT = Font(
        size=12,
        bold=True,
        color='C0392B',  # 深红色，突出总计
        name='微软雅黑',
    )

    # 标签字样（采购方、日期等标签）
    LABEL_FONT = Font(
        size=11,
        bold=True,
        color='7F8C8D',  # 中灰色
        name='微软雅黑',
    )

    # ========== 对齐方式 ==========
    # 标题居中
    TITLE_ALIGNMENT = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=False,
    )

    # 表头居中
    HEADER_ALIGNMENT = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True,
    )

    # 数据左对齐（文本）
    DATA_LEFT_ALIGNMENT = Alignment(
        horizontal='left',
        vertical='center',
        indent=1,
    )

    # 数据居中（序号、单位等）
    DATA_CENTER_ALIGNMENT = Alignment(
        horizontal='center',
        vertical='center',
    )

    # 数据右对齐（数值）
    DATA_RIGHT_ALIGNMENT = Alignment(
        horizontal='right',
        vertical='center',
    )

    # ========== 背景色填充 ==========
    # 表头背景（深蓝色系，专业感）
    HEADER_FILL = PatternFill(
        start_color='3498DB',  # 专业蓝
        end_color='3498DB',
        fill_type='solid',
    )

    # 总计行背景（浅黄色，突出但不刺眼）
    TOTAL_FILL = PatternFill(
        start_color='FFF9E6',  # 浅黄色
        end_color='FFF9E6',
        fill_type='solid',
    )

    # 第一列背景（序号列，浅灰色区分）
    SEQUENCE_FILL = PatternFill(
        start_color='F8F9FA',  # 极浅灰色
        end_color='F8F9FA',
        fill_type='solid',
    )

    # ========== 数字格式 ==========
    # 金额格式（带人民币符号、千分位）
    CURRENCY_FORMAT = '¥#,##0.00'

    # 数量格式（带千分位）
    NUMBER_FORMAT = '#,##0.0'

    # ========== 行高定义 ==========
    ROW_HEIGHT_TITLE = 40       # 标题行
    ROW_HEIGHT_INFO = 25        # 信息行
    ROW_HEIGHT_SPACER = 12      # 分隔空行
    ROW_HEIGHT_HEADER = 30      # 表头行
    ROW_HEIGHT_DATA = 24        # 数据行
    ROW_HEIGHT_TOTAL = 32       # 总计行

    # ========== 列宽定义 ==========
    COLUMN_WIDTHS = {
        "A": 8,    # 序号
        "B": 35,   # 商品名称
        "C": 20,   # 规格型号
        "D": 10,   # 单位
        "E": 14,   # 数量
        "F": 16,   # 单价
        "G": 16,   # 金额
        "H": 28,   # 备注
    }

from . import PurchaseReceipt, PurchaseItem


class ExcelHandler:
    """
    Excel处理器

    负责读取和写入采购明细Excel文件。
    """

    # 表格列定义
    COLS = {
        "sequence": "A",    # 序号
        "name": "B",        # 商品名称
        "spec": "C",        # 规格型号
        "unit": "D",        # 单位
        "quantity": "E",    # 采购数量
        "unit_price": "F",  # 单价（元）
        "amount": "G",      # 金额（元）
        "remark": "H",      # 备注
    }

    # 固定值
    DEFAULT_PURCHASER = "梁程程妈妈"
    DEFAULT_PAYMENT_METHOD = "转账"

    def __init__(self, file_path: Path, auto_backup: bool = True):
        """
        初始化Excel处理器

        Args:
            file_path: Excel文件路径
            auto_backup: 是否自动备份
        """
        self.file_path = Path(file_path).expanduser()
        self.auto_backup = auto_backup
        self._workbook = None

    def _load_workbook(self) -> Workbook:
        """加载工作簿"""
        if self._workbook is None:
            if self.file_path.exists():
                self._workbook = load_workbook(self.file_path)
            else:
                # 创建新工作簿
                self._workbook = Workbook()
                # 删除默认Sheet
                if "Sheet" in self._workbook.sheetnames:
                    self._workbook.remove(self._workbook["Sheet"])
        return self._workbook

    def _backup(self):
        """备份Excel文件"""
        if self.auto_backup and self.file_path.exists():
            backup_path = self.file_path.with_suffix(
                f".bak.{date.today().isoformat()}"
            )
            shutil.copy2(self.file_path, backup_path)

    def sheet_exists(self, sheet_name: str) -> bool:
        """
        检查Sheet是否存在

        Args:
            sheet_name: Sheet名称

        Returns:
            是否存在
        """
        wb = self._load_workbook()
        return sheet_name in wb.sheetnames

    def add_receipt(self, receipt: PurchaseReceipt) -> None:
        """
        添加收据到Excel

        Args:
            receipt: 收据数据
        """
        # 备份
        self._backup()

        # 加载工作簿
        wb = self._load_workbook()

        # 创建或更新Sheet
        sheet_name = receipt.sheet_name
        if sheet_name in wb.sheetnames:
            # 更新现有Sheet
            ws = wb[sheet_name]
            self._update_sheet(ws, receipt)
        else:
            # 创建新Sheet
            ws = wb.create_sheet(title=sheet_name)
            self._create_sheet(ws, receipt)

        # 保存
        self._save()

    def _create_sheet(self, ws, receipt: PurchaseReceipt):
        """
        创建新Sheet - 业界最佳实践布局

        布局设计：
        第1行: 收据标题（深色文字，居中，大字体）
        第2行: 采购方信息（分列布局，标签加粗）
        第3行: 空行（视觉分隔）
        第4行: 表头（深蓝色背景，白色文字）
        第5行起: 商品明细（带边框，序号列浅灰背景）
        倒数第2行: 总计（浅黄背景，深红文字，双线下边框）
        最后1行: 交付信息（分列显示）

        样式特点：
        - 使用微软雅黑字体，更清晰易读
        - 表头深蓝色背景，专业感强
        - 总计行突出显示，便于快速查看
        - 数值右对齐，文本左对齐，符合阅读习惯
        """
        # ========== 1. 标题区域 ==========
        ws.merge_cells("A1:H1")
        ws["A1"] = receipt.title
        ws["A1"].font = Styles.TITLE_FONT
        ws["A1"].alignment = Styles.TITLE_ALIGNMENT
        ws.row_dimensions[1].height = Styles.ROW_HEIGHT_TITLE

        # ========== 2. 采购方信息（分列布局）==========
        ws["A2"] = "采购方："
        ws["B2"] = receipt.purchaser
        ws["D2"] = "日期："
        ws["E2"] = receipt.delivery_date.strftime("%Y年%m月%d日")

        # 标签样式
        ws["A2"].font = Styles.LABEL_FONT
        ws["A2"].alignment = Styles.DATA_LEFT_ALIGNMENT

        ws["D2"].font = Styles.LABEL_FONT
        ws["D2"].alignment = Styles.DATA_LEFT_ALIGNMENT

        # 内容样式
        ws["B2"].font = Styles.DATA_FONT
        ws["B2"].alignment = Styles.DATA_LEFT_ALIGNMENT

        ws["E2"].font = Styles.DATA_FONT
        ws["E2"].alignment = Styles.DATA_LEFT_ALIGNMENT

        ws.row_dimensions[2].height = Styles.ROW_HEIGHT_INFO

        # ========== 3. 空行分隔 ==========
        ws.row_dimensions[3].height = Styles.ROW_HEIGHT_SPACER

        # ========== 4. 表头行（深蓝色背景）==========
        headers = ["序号", "商品名称", "规格型号", "单位", "数量", "单价（元）", "金额（元）", "备注"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = Styles.HEADER_FONT
            cell.alignment = Styles.HEADER_ALIGNMENT
            cell.fill = Styles.HEADER_FILL
            cell.border = Styles.HEADER_BORDER

        ws.row_dimensions[4].height = Styles.ROW_HEIGHT_HEADER

        # ========== 5. 商品明细 ==========
        row = 5
        for item in receipt.items:
            # 序号 - 居中，浅灰背景
            ws[f"A{row}"] = item.sequence
            ws[f"A{row}"].font = Styles.DATA_FONT
            ws[f"A{row}"].alignment = Styles.DATA_CENTER_ALIGNMENT
            ws[f"A{row}"].fill = Styles.SEQUENCE_FILL
            ws[f"A{row}"].border = Styles.THIN_BORDER

            # 商品名称 - 左对齐
            ws[f"B{row}"] = item.name
            ws[f"B{row}"].font = Styles.DATA_FONT
            ws[f"B{row}"].alignment = Styles.DATA_LEFT_ALIGNMENT
            ws[f"B{row}"].border = Styles.THIN_BORDER

            # 规格型号 - 左对齐
            ws[f"C{row}"] = item.spec or ""
            ws[f"C{row}"].font = Styles.DATA_FONT
            ws[f"C{row}"].alignment = Styles.DATA_LEFT_ALIGNMENT
            ws[f"C{row}"].border = Styles.THIN_BORDER

            # 单位 - 居中，浅灰背景
            ws[f"D{row}"] = item.unit
            ws[f"D{row}"].font = Styles.DATA_FONT
            ws[f"D{row}"].alignment = Styles.DATA_CENTER_ALIGNMENT
            ws[f"D{row}"].fill = Styles.SEQUENCE_FILL
            ws[f"D{row}"].border = Styles.THIN_BORDER

            # 数量 - 右对齐
            ws[f"E{row}"] = float(item.quantity)
            ws[f"E{row}"].font = Styles.DATA_FONT
            ws[f"E{row}"].number_format = Styles.NUMBER_FORMAT
            ws[f"E{row}"].alignment = Styles.DATA_RIGHT_ALIGNMENT
            ws[f"E{row}"].border = Styles.THIN_BORDER

            # 单价 - 右对齐，货币格式
            ws[f"F{row}"] = float(item.unit_price)
            ws[f"F{row}"].font = Styles.DATA_FONT
            ws[f"F{row}"].number_format = Styles.CURRENCY_FORMAT
            ws[f"F{row}"].alignment = Styles.DATA_RIGHT_ALIGNMENT
            ws[f"F{row}"].border = Styles.THIN_BORDER

            # 金额 - 右对齐，公式+货币格式
            ws[f"G{row}"] = f"=E{row}*F{row}"
            ws[f"G{row}"].font = Styles.DATA_FONT
            ws[f"G{row}"].number_format = Styles.CURRENCY_FORMAT
            ws[f"G{row}"].alignment = Styles.DATA_RIGHT_ALIGNMENT
            ws[f"G{row}"].border = Styles.THIN_BORDER

            # 备注 - 左对齐
            ws[f"H{row}"] = item.remark or ""
            ws[f"H{row}"].font = Styles.DATA_FONT
            ws[f"H{row}"].alignment = Styles.DATA_LEFT_ALIGNMENT
            ws[f"H{row}"].border = Styles.THIN_BORDER

            ws.row_dimensions[row].height = Styles.ROW_HEIGHT_DATA
            row += 1

        # ========== 6. 总计行（浅黄背景，突出显示）==========
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = "总    计"
        ws[f"A{row}"].font = Styles.TOTAL_FONT
        ws[f"A{row}"].alignment = Styles.DATA_CENTER_ALIGNMENT
        ws[f"A{row}"].fill = Styles.TOTAL_FILL
        ws[f"A{row}"].border = Styles.TOTAL_BORDER

        ws[f"G{row}"] = f"=SUM(G5:G{row-1})"
        ws[f"G{row}"].font = Styles.TOTAL_FONT
        ws[f"G{row}"].number_format = Styles.CURRENCY_FORMAT
        ws[f"G{row}"].alignment = Styles.DATA_RIGHT_ALIGNMENT
        ws[f"G{row}"].fill = Styles.TOTAL_FILL
        ws[f"G{row}"].border = Styles.TOTAL_BORDER

        ws[f"H{row}"].fill = Styles.TOTAL_FILL
        ws[f"H{row}"].border = Styles.TOTAL_BORDER

        ws.row_dimensions[row].height = Styles.ROW_HEIGHT_TOTAL

        # ========== 7. 交付信息（分列）==========
        row += 1
        ws[f"A{row}"] = "交付日期："
        ws[f"B{row}"] = receipt.delivery_date.strftime("%Y年%m月%d日")
        ws[f"D{row}"] = "付款方式："
        ws[f"E{row}"] = receipt.payment_method

        # 标签样式
        ws[f"A{row}"].font = Styles.LABEL_FONT
        ws[f"A{row}"].alignment = Styles.DATA_LEFT_ALIGNMENT

        ws[f"D{row}"].font = Styles.LABEL_FONT
        ws[f"D{row}"].alignment = Styles.DATA_LEFT_ALIGNMENT

        # 内容样式
        ws[f"B{row}"].font = Styles.DATA_FONT
        ws[f"B{row}"].alignment = Styles.DATA_LEFT_ALIGNMENT

        ws[f"E{row}"].font = Styles.DATA_FONT
        ws[f"E{row}"].alignment = Styles.DATA_LEFT_ALIGNMENT

        ws.row_dimensions[row].height = Styles.ROW_HEIGHT_INFO

        # ========== 8. 插入原始收据图片（表格下方）==========
        image_row = row + 2  # 在交付信息后空一行

        if receipt.source_file:
            image_path = Path(receipt.source_file)
            if image_path.exists():
                try:
                    # 加载图片
                    img = Image(str(image_path))

                    # 设置图片大小（保持比例，限制最大宽度）
                    max_width = 500  # 像素，加宽以适应表格宽度
                    if img.width > max_width:
                        ratio = max_width / img.width
                        img.width = max_width
                        img.height = int(img.height * ratio)

                    # 将图片放在表格下方
                    img.anchor = f"A{image_row}"
                    ws.add_image(img)

                    # 添加"原始凭证"标题
                    ws[f"A{image_row}"] = "原始凭证："
                    ws[f"A{image_row}"].font = Styles.LABEL_FONT
                    ws[f"A{image_row}"].alignment = Styles.DATA_LEFT_ALIGNMENT

                    # 设置图片所在行的高度
                    ws.row_dimensions[image_row].height = 20  # 标题行高度
                    image_row += 1
                    # 预留图片空间（每行约15像素，动态计算）
                    img_height_rows = max(15, int(img.height / 15) + 2)
                    for r in range(image_row, image_row + img_height_rows):
                        if r not in ws.row_dimensions:
                            ws.row_dimensions[r].height = 15

                    logger.info(f"✓ 已插入收据图片: {image_path.name}")
                except Exception as e:
                    logger.warning(f"插入图片失败: {e}")

        # ========== 9. 设置列宽 ==========
        for col, width in Styles.COLUMN_WIDTHS.items():
            ws.column_dimensions[col].width = width

    def _update_sheet(self, ws, receipt: PurchaseReceipt):
        """
        更新现有Sheet
        """
        # 清空现有数据（保留前4行）
        max_row = ws.max_row
        if max_row > 4:
            ws.delete_rows(5, max_row - 4)

        # 重新创建
        self._create_sheet(ws, receipt)

    def _save(self):
        """保存工作簿"""
        self.file_path.parent.mkdir(parents=True, exist_ok=True)
        if self._workbook:
            self._workbook.save(self.file_path)

    def read_receipt(self, sheet_name: str) -> Optional[PurchaseReceipt]:
        """
        从Sheet读取收据

        Args:
            sheet_name: Sheet名称

        Returns:
            收据数据，如果Sheet不存在返回None
        """
        wb = self._load_workbook()

        if sheet_name not in wb.sheetnames:
            return None

        ws = wb[sheet_name]

        # 1. 读取标题
        title = ws["A1"].value or ""

        # 2. 读取采购方
        purchaser = ws["B2"].value or self.DEFAULT_PURCHASER

        # 3. 获取最后一行
        last_row = ws.max_row

        # 4. 读取交付日期（支持多种格式）
        # 首先尝试从 E2 读取（日期字段）
        delivery_date = self._parse_date_from_cell(ws["E2"].value)

        # 如果 E2 无效，尝试从最后一行读取（交付日期字段）
        if delivery_date is None:
            # 查找包含"交付日期"的行
            for row in range(last_row, max(1, last_row - 5), -1):
                a_val = ws[f"A{row}"].value
                b_val = ws[f"B{row}"].value
                if a_val and "交付日期" in str(a_val):
                    delivery_date = self._parse_date_from_cell(b_val)
                    break
                elif b_val and "交付日期" in str(b_val):
                    delivery_date = self._parse_date_from_cell(a_val)
                    break

        # 如果仍然无效，使用今天
        if delivery_date is None:
            delivery_date = date.today()

        # 5. 读取商品明细
        items = []
        row = 5
        while row <= last_row:
            seq = ws[f"A{row}"].value
            name = ws[f"B{row}"].value

            # 跳过空行
            if not seq and not name:
                row += 1
                continue

            # 跳过"总计金额"行和交付信息行
            if isinstance(seq, str) and ("总计" in seq or "交付" in seq or "付款" in seq):
                break
            if isinstance(name, str) and ("总计" in name or "交付" in name or "付款" in name):
                break

            # 确保seq是数字
            try:
                seq_num = int(seq) if seq else 0
            except (ValueError, TypeError):
                row += 1
                continue

            item = PurchaseItem(
                sequence=seq_num,
                name=name or "",
                spec=ws[f"C{row}"].value,
                unit=ws[f"D{row}"].value or "个",
                quantity=Decimal(str(ws[f"E{row}"].value or 0)),
                unit_price=Decimal(str(ws[f"F{row}"].value or 0)),
                remark=ws[f"H{row}"].value,
            )
            items.append(item)
            row += 1

        return PurchaseReceipt(
            title=title,
            delivery_date=delivery_date,
            purchaser=purchaser,
            items=items,
            recognition_method="excel",
        )

    def list_sheets(self) -> List[str]:
        """
        列出所有Sheet名称

        Returns:
            Sheet名称列表
        """
        wb = self._load_workbook()
        return wb.sheetnames

    def _parse_date_from_cell(self, cell_value) -> Optional[date]:
        """
        从单元格值解析日期（支持多种格式）

        支持的格式：
        - Excel 日期序列号（整数，如 45980）
        - 字符串格式：2025-01-20
        - 字符串格式：2025年01月20日
        - datetime 对象

        Args:
            cell_value: 单元格值

        Returns:
            日期对象，解析失败返回 None
        """
        if cell_value is None:
            return None

        # Excel 日期序列号（整数 > 30000）
        if isinstance(cell_value, (int, float)) and cell_value > 30000:
            try:
                excel_date = datetime(1899, 12, 30) + timedelta(days=cell_value)
                return excel_date.date()
            except (ValueError, OverflowError):
                pass

        # datetime 对象
        if isinstance(cell_value, date) and not isinstance(cell_value, datetime):
            return cell_value

        if isinstance(cell_value, datetime):
            return cell_value.date()

        # 字符串格式
        if isinstance(cell_value, str):
            # 2025-01-20 格式
            match = re.search(r'(\d{4})-(\d{1,2})-(\d{1,2})', cell_value)
            if match:
                year, month, day = match.groups()
                try:
                    return date(int(year), int(month), int(day))
                except ValueError:
                    pass

            # 2025年01月20日 格式
            match = re.search(r'(\d{4})年(\d{1,2})月(\d{1,2})日', cell_value)
            if match:
                year, month, day = match.groups()
                try:
                    return date(int(year), int(month), int(day))
                except ValueError:
                    pass

        return None

    def _parse_delivery_date(self, delivery_info: str) -> date:
        """
        从交付信息中解析日期

        格式：交付日期：2025-1-20    付款方式：转账

        Args:
            delivery_info: 交付信息字符串

        Returns:
            日期
        """
        # 提取日期部分
        match = re.search(r'(\d{4})-(\d{1,2})-(\d{1,2})', delivery_info)
        if match:
            year, month, day = match.groups()
            return date(int(year), int(month), int(day))

        # 默认返回今天
        return date.today()

    def get_statistics(self) -> dict:
        """
        获取统计信息

        Returns:
            统计信息字典
        """
        wb = self._load_workbook()
        sheets = wb.sheetnames

        receipts = []
        for sheet_name in sheets:
            receipt = self.read_receipt(sheet_name)
            if receipt:
                receipts.append(receipt)

        total_amount = sum(r.total_amount for r in receipts)
        total_items = sum(r.item_count for r in receipts)

        return {
            "sheet_count": len(sheets),
            "receipt_count": len(receipts),
            "total_amount": float(total_amount),
            "total_items": total_items,
            "recent_receipts": [
                {
                    "title": r.title,
                    "date": r.delivery_date.isoformat(),
                    "amount": float(r.total_amount),
                    "sheet": r.sheet_name,
                }
                for r in sorted(receipts, key=lambda x: x.delivery_date, reverse=True)[:10]
            ],
        }

    def close(self):
        """关闭工作簿"""
        if self._workbook:
            self._workbook.close()
            self._workbook = None

    def beautify_all_sheets(self) -> int:
        """
        美化所有Sheet - 应用最新的样式到现有收据

        这个方法会读取每个Sheet的收据数据，然后用新样式重新创建。
        适用于批量更新已有收据的样式。

        Returns:
            美化的Sheet数量
        """
        # 备份
        self._backup()

        # 加载工作簿
        wb = self._load_workbook()
        sheet_names = wb.sheetnames[:]

        beautified_count = 0

        for sheet_name in sheet_names:
            # 读取现有收据数据
            receipt = self.read_receipt(sheet_name)
            if receipt:
                # 获取原始Sheet
                ws = wb[sheet_name]

                # 清空现有内容
                max_row = ws.max_row
                if max_row > 0:
                    ws.delete_rows(1, max_row)

                # 用新样式重新创建
                self._create_sheet(ws, receipt)
                beautified_count += 1

        # 保存
        self._save()

        return beautified_count


# 便捷函数
def load_excel(file_path: str) -> ExcelHandler:
    """
    加载Excel文件

    Args:
        file_path: 文件路径

    Returns:
        ExcelHandler实例
    """
    return ExcelHandler(Path(file_path))


def create_receipt_from_excel(file_path: str, sheet_name: str) -> Optional[PurchaseReceipt]:
    """
    从Excel创建收据

    Args:
        file_path: Excel文件路径
        sheet_name: Sheet名称

    Returns:
        收据数据
    """
    handler = load_excel(file_path)
    receipt = handler.read_receipt(sheet_name)
    handler.close()
    return receipt


def save_receipt_to_excel(file_path: str, receipt: PurchaseReceipt) -> None:
    """
    保存收据到Excel

    Args:
        file_path: Excel文件路径
        receipt: 收据数据
    """
    handler = load_excel(file_path)
    handler.add_receipt(receipt)
    handler.close()
